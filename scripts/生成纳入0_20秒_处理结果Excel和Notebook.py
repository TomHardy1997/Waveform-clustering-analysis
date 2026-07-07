from pathlib import Path
import json

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


PROJECT_ROOT = Path(__file__).resolve().parents[1]
VIS_DIR = PROJECT_ROOT / "results" / "visualization"
RESULT_DIR = PROJECT_ROOT / "results" / "preprocessing"
CSV_DIR = RESULT_DIR / "单文件CSV数据"
SINGLE_XLSX_DIR = RESULT_DIR / "单文件独立处理结果"
SUMMARY_XLSX = RESULT_DIR / "纳入0-20秒_批量预处理特征汇总.xlsx"
NOTEBOOK = RESULT_DIR / "纳入0-20秒_预处理特征可视化.ipynb"
README = RESULT_DIR / "纳入0-20秒_预处理特征结果说明.md"


def style_workbook(path: Path):
    wb = load_workbook(path)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    zebra = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = center
                cell.border = border
                if cell.row % 2 == 0:
                    cell.fill = zebra
        for col in ws.columns:
            letter = col[0].column_letter
            max_len = 8
            for cell in col[:200]:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[letter].width = min(max_len + 2, 34)
    wb.save(path)


def make_single_excels(raw_long, preprocess, peaks, valleys, stages, summary):
    SINGLE_XLSX_DIR.mkdir(parents=True, exist_ok=True)
    for i, (file_id, one_pre) in enumerate(preprocess.groupby("file_id", sort=True), start=1):
        row = summary[summary["file_id"] == file_id].iloc[0]
        filename = row["文件名"]
        out = SINGLE_XLSX_DIR / f"处理结果_{filename}"

        raw = raw_long[raw_long["file_id"] == file_id].copy()
        raw_sheet = pd.DataFrame(
            {
                "Time_F_F0": raw["time"],
                "F_F0": raw["f_f0"],
                "Time_dF_F0": raw["time"],
                "dF_F0": raw["dff"],
            }
        )
        pre_sheet = one_pre.drop(columns=["file_id", "文件名"], errors="ignore")
        peak_sheet = peaks[peaks["file_id"] == file_id].drop(columns=["file_id", "文件名"], errors="ignore")
        valley_sheet = valleys[valleys["file_id"] == file_id].drop(columns=["file_id", "文件名"], errors="ignore")
        stage_sheet = stages[stages["file_id"] == file_id].drop(columns=["file_id", "文件名"], errors="ignore")

        stat_pairs = row.drop(labels=["file_id"]).to_frame()
        stat_pairs = stat_pairs.reset_index()
        stat_pairs.columns = ["指标名称", "指标值"]

        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            raw_sheet.to_excel(writer, sheet_name="原始数据", index=False)
            pre_sheet.to_excel(writer, sheet_name="预处理后数据", index=False)
            peak_sheet.to_excel(writer, sheet_name="峰值特征", index=False)
            valley_sheet.to_excel(writer, sheet_name="波谷特征", index=False)
            stage_sheet.to_excel(writer, sheet_name="分阶段特征", index=False)
            stat_pairs.to_excel(writer, sheet_name="整体统计特征", index=False)
        style_workbook(out)
        if i % 25 == 0 or i == summary["file_id"].nunique():
            print(f"已生成 {i}/{summary['file_id'].nunique()} 个单文件处理结果Excel")


def make_summary_workbook(raw_long, preprocess, peaks, valleys, stages, summary, core):
    feature_desc = pd.DataFrame(
        [
            ("总波峰数量", "0-20秒内检测到的上升峰数量"),
            ("总波谷数量", "0-20秒内检测到的下降谷数量"),
            ("最大波峰幅度", "相对基线的最大正向ΔF/F0"),
            ("最大波峰出现时间", "最大上升峰出现的秒数"),
            ("最大波谷幅度", "相对基线的最大负向ΔF/F0"),
            ("最大波谷出现时间", "最大下降谷出现的秒数"),
            ("ON反应幅度", "5-10秒刺激期最大上升峰幅度"),
            ("ON反应潜伏期", "从5秒刺激开始到刺激期第一个峰的时间"),
            ("OFF反应幅度", "10-20秒撤光后最大上升峰幅度"),
            ("OFF反应潜伏期", "从10秒撤光开始到撤光后第一个峰的时间"),
            ("刺激期相对刺激前变化量", "刺激期平均值减去刺激前平均值"),
            ("撤光后相对刺激前变化量", "撤光后平均值减去刺激前平均值"),
            ("刺激期AUC", "5-10秒阶段曲线面积"),
            ("撤光后AUC", "10-20秒阶段曲线面积"),
            ("恢复斜率", "10-20秒撤光后阶段的线性斜率"),
            ("反应类型初判", "根据ON/OFF幅度和刺激期变化得到的粗略类型"),
        ],
        columns=["特征名", "解释"],
    )

    overview = pd.DataFrame(
        {
            "项目": [
                "处理文件数",
                "总波峰数量均值",
                "总波谷数量均值",
                "ON反应幅度均值",
                "刺激期相对刺激前变化量均值",
                "撤光后相对刺激前变化量均值",
            ],
            "数值": [
                len(summary),
                summary["总波峰数量"].mean(),
                summary["总波谷数量"].mean(),
                summary["ON反应幅度"].mean(),
                summary["刺激期相对刺激前变化量"].mean(),
                summary["撤光后相对刺激前变化量"].mean(),
            ],
        }
    )

    type_count = summary["反应类型初判"].value_counts().rename_axis("反应类型初判").reset_index(name="数量")

    with pd.ExcelWriter(SUMMARY_XLSX, engine="openpyxl") as writer:
        overview.to_excel(writer, sheet_name="总览", index=False)
        type_count.to_excel(writer, sheet_name="反应类型统计", index=False)
        summary.to_excel(writer, sheet_name="整体统计和全部特征", index=False)
        core.to_excel(writer, sheet_name="核心聚类特征", index=False)
        stages.to_excel(writer, sheet_name="分阶段特征", index=False)
        peaks.to_excel(writer, sheet_name="全部峰值特征", index=False)
        valleys.to_excel(writer, sheet_name="全部波谷特征", index=False)
        feature_desc.to_excel(writer, sheet_name="特征说明", index=False)
    style_workbook(SUMMARY_XLSX)


def make_notebook():
    cells = []

    def md(text):
        cells.append({"cell_type": "markdown", "metadata": {}, "source": text.splitlines(True)})

    def code(text):
        cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "outputs": [], "source": text.splitlines(True)})

    md("""# 纳入 0-20 秒 Excel 的预处理、特征提取和可视化\n\n这份 notebook 用来查看已经生成的结果。处理流程参考 `Processing_waveform.py` 和示例处理结果 Excel，并加入 0-5、5-10、10-20 秒阶段特征。\n""")
    code("""from pathlib import Path\nimport pandas as pd\nfrom IPython.display import Image, display\n\nROOT = Path('/Users/tangdi/Desktop/波形Excel数据/整理后_0-20秒原始Excel')\nRESULT_DIR = ROOT / '批量预处理特征结果'\nCSV_DIR = RESULT_DIR / '单文件CSV数据'\nPLOT_DIR = RESULT_DIR / '单文件预处理波形图'\nOVERVIEW_DIR = RESULT_DIR / '总览图'\nprint(RESULT_DIR)\n""")
    md("""## 1. 读取总特征表""")
    code("""summary = pd.read_csv(RESULT_DIR / '全部文件_整体统计和核心特征.csv')\ncore = pd.read_csv(RESULT_DIR / '用于聚类的核心特征表.csv')\ndisplay(summary.head())\ndisplay(core.head())\n""")
    md("""## 2. 查看关键统计""")
    code("""display(summary['反应类型初判'].value_counts().rename_axis('反应类型').reset_index(name='数量'))\ncols = ['总波峰数量','总波谷数量','ON反应幅度','刺激期相对刺激前变化量','撤光后相对刺激前变化量','波形标准差']\ndisplay(summary[cols].describe())\n""")
    md("""## 3. 查看总览图""")
    code("""for name in ['全部预处理波形叠加.png', '预处理波形热图_按刺激响应排序.png', '核心特征分布.png']:\n    display(Image(filename=str(OVERVIEW_DIR / name)))\n""")
    md("""## 4. 查看单个细胞的处理结果""")
    code("""example = summary.iloc[0]['file_id']\nprint('示例:', example)\ndisplay(Image(filename=str(PLOT_DIR / f'{example}_预处理峰谷图.png')))\n""")
    md("""## 5. 输出文件位置\n\n- 单文件 Excel：`批量预处理特征结果/单文件独立处理结果`\n- 汇总 Excel：`批量预处理特征结果/纳入0-20秒_批量预处理特征汇总.xlsx`\n- 单文件图：`批量预处理特征结果/单文件预处理波形图`\n- 总览图：`批量预处理特征结果/总览图`\n""")

    nb = {
        "cells": cells,
        "metadata": {
            "kernelspec": {"display_name": "Python 3", "language": "python", "name": "python3"},
            "language_info": {"name": "python", "pygments_lexer": "ipython3"},
        },
        "nbformat": 4,
        "nbformat_minor": 5,
    }
    NOTEBOOK.write_text(json.dumps(nb, ensure_ascii=False, indent=2), encoding="utf-8")


def make_readme(summary):
    README.write_text(
        f"""# 纳入 0-20 秒 Excel 批量预处理和特征提取结果\n\n## 输入\n\n输入数据来自：\n\n`/Users/tangdi/Desktop/波形Excel数据/整理后_0-20秒原始Excel/纳入_0-20秒Excel`\n\n处理文件数：`{len(summary)}` 个。\n\n## 处理内容\n\n每个文件按照示例 `处理结果_113520-2.xlsx` 的形式生成处理结果，包括：\n\n- 原始数据\n- 预处理后数据\n- 峰值特征\n- 波谷特征\n- 分阶段特征\n- 整体统计特征\n\n## 时间段定义\n\n- 刺激前：0-5 秒\n- 刺激期：5-10 秒\n- 撤光后：10-20 秒\n\n## 输出\n\n- `单文件独立处理结果`：每个细胞一个 Excel\n- `单文件预处理波形图`：每个细胞一张预处理和峰谷标记图\n- `总览图`：全部波形叠加、热图、特征分布\n- `纳入0-20秒_批量预处理特征汇总.xlsx`：总特征表\n- `纳入0-20秒_预处理特征可视化.ipynb`：查看结果的 notebook\n\n## 说明\n\n这版处理参考前人脚本的预处理步骤：异常值截断、基线校正、Savitzky-Golay 平滑、峰值检测。额外增加了波谷检测、分阶段特征、ON/OFF 特征和用于聚类的核心特征。\n""",
        encoding="utf-8",
    )


def main():
    raw_long = pd.read_csv(VIS_DIR / "纳入文件_0-20秒原始波形长表.csv")
    preprocess = pd.read_csv(CSV_DIR / "全部文件_预处理后数据.csv")
    peaks = pd.read_csv(CSV_DIR / "全部文件_峰值特征.csv")
    valleys = pd.read_csv(CSV_DIR / "全部文件_波谷特征.csv")
    stages = pd.read_csv(CSV_DIR / "全部文件_分阶段特征.csv")
    summary = pd.read_csv(RESULT_DIR / "全部文件_整体统计和核心特征.csv")
    core = pd.read_csv(RESULT_DIR / "用于聚类的核心特征表.csv")

    make_single_excels(raw_long, preprocess, peaks, valleys, stages, summary)
    make_summary_workbook(raw_long, preprocess, peaks, valleys, stages, summary, core)
    make_notebook()
    make_readme(summary)

    print(f"单文件Excel: {SINGLE_XLSX_DIR}")
    print(f"汇总Excel: {SUMMARY_XLSX}")
    print(f"Notebook: {NOTEBOOK}")
    print(f"说明文档: {README}")


if __name__ == "__main__":
    main()
