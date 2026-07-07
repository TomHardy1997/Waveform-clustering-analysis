# ===================== 1. 导入依赖库 =====================
import pandas as pd
import numpy as np
from scipy.signal import savgol_filter, find_peaks, peak_widths
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from tqdm import tqdm  # 进度条库，直观显示处理进度

# ===================== 2. 核心配置参数（仅需修改这里即可） =====================
# 【必填】原始数据所在文件夹路径（所有待处理的xlsx文件都放在这个文件夹里）
input_folder = "E:\\YKYJS\\视网膜钙离子成像\\波形Excel数据(1)\\波形Excel数据"  # 替换为你的文件夹路径，比如"/mnt/钙成像数据"

# 【可选】输出结果文件夹路径（自动创建，无需手动新建）
output_folder = "E:\\YKYJS\\视网膜钙离子成像\\波形Excel数据(1)\\波形Excel数据\\result"

# 【可选】是否生成每个文件的波形图（True=生成，False=不生成，节省时间）
generate_plot = True

# 【可选】文件匹配规则：只处理文件名包含指定关键词的xlsx文件，空值=处理所有xlsx
file_keyword = ""  # 比如"钙成像"，只处理文件名带"钙成像"的文件

# 【处理参数】和之前单文件完全一致，所有文件统一使用该参数
savgol_window = 51  # 平滑滤波器窗口大小（奇数，越大平滑效果越强）
savgol_polyorder = 3  # 滤波器多项式阶数
peak_prominence = 0.002  # 峰值突出度（越小越容易检测到低幅度峰）
peak_height = 0.005  # 峰值最小高度（过滤掉过低的噪声峰）
baseline_ratio = 0.1  # 基线计算比例（前10%数据作为基线）

# 【列名配置】和你的原始表格列名完全匹配，和之前单文件一致
raw_columns = ['Time - F/F0', 'F/F0 - F/F0', 'Time - dF/F0', 'F/F0 - dF/F0']
# 处理后的统一列名
rename_columns = ['Time_F_F0', 'F_F0', 'Time_dF_F0', 'dF_F0']

# ===================== 3. 初始化输出文件夹 =====================
# 自动创建输出文件夹
os.makedirs(output_folder, exist_ok=True)
# 单文件处理结果子文件夹
single_file_output_folder = os.path.join(output_folder, "单文件独立处理结果")
os.makedirs(single_file_output_folder, exist_ok=True)
# 波形图子文件夹
if generate_plot:
    plot_output_folder = os.path.join(output_folder, "单文件波形图")
    os.makedirs(plot_output_folder, exist_ok=True)

# 汇总结果列表
summary_result_list = []

# 解决matplotlib中文显示
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False

# 定义Excel美化样式（和之前单文件完全一致）
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
center_alignment = Alignment(horizontal='center', vertical='center')
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
zebra_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

# ===================== 4. 定义单文件处理函数（核心逻辑，和之前完全一致） =====================
def process_single_file(file_path, file_name):
    try:
        # 4.1 读取原始数据
        df_raw = pd.read_excel(file_path)
        # 校验列名是否匹配
        if not all(col in df_raw.columns for col in raw_columns):
            print(f"⚠️  文件 {file_name} 列名不匹配，跳过处理")
            return None
        # 重命名列名
        df_raw = df_raw[raw_columns].rename(columns=dict(zip(raw_columns, rename_columns)))
        # 提取核心数据
        time = df_raw['Time_F_F0'].values
        f_f0_raw = df_raw['F_F0'].values
        df_f0_raw = df_raw['dF_F0'].values
        sample_rate = 1 / np.mean(np.diff(time))  # 计算采样频率

        # 4.2 数据预处理
        # 异常值处理（3倍标准差法截断）
        mean_f = np.mean(f_f0_raw)
        std_f = np.std(f_f0_raw)
        f_f0_clip = np.clip(f_f0_raw, mean_f - 3*std_f, mean_f + 3*std_f)

        # 基线校正（以前10%数据的均值作为F0基线）
        baseline_length = int(len(f_f0_clip) * baseline_ratio)
        f0_baseline = np.mean(f_f0_clip[:baseline_length])
        f_f0_baseline = f_f0_clip / f0_baseline  # 校正后基线稳定在1附近

        # 平滑去噪（Savitzky-Golay滤波器）
        f_f0_smooth = savgol_filter(f_f0_baseline, window_length=savgol_window, polyorder=savgol_polyorder)
        df_f0_smooth = savgol_filter(df_f0_raw, window_length=savgol_window, polyorder=savgol_polyorder)

        # 组装预处理后的数据
        df_preprocess = pd.DataFrame({
            '时间(秒)': time,
            '原始F/F0': f_f0_raw,
            '异常值截断后F/F0': f_f0_clip,
            '基线校正后F/F0': f_f0_baseline,
            '平滑去噪后F/F0': f_f0_smooth,
            '原始dF/F0': df_f0_raw,
            '平滑去噪后dF/F0': df_f0_smooth
        })

        # 4.3 钙峰检测与特征提取
        # 钙峰检测
        peaks, peak_props = find_peaks(
            f_f0_smooth,
            height=peak_height,
            prominence=peak_prominence,
            distance=int(sample_rate * 0.5)  # 两个峰最小间隔0.5秒
        )
        peak_count = len(peaks)

        # 单峰特征提取
        peak_features = []
        for i, peak_idx in enumerate(peaks):
            # 基础信息
            peak_time = time[peak_idx]
            peak_amplitude = f_f0_smooth[peak_idx] - 1  # ΔF/F0
            # 半高宽计算
            widths, width_heights, left_ips, right_ips = peak_widths(
                f_f0_smooth, [peak_idx], rel_height=0.5
            )
            fwhm = widths[0] / sample_rate
            fwhm_left = left_ips[0] / sample_rate
            fwhm_right = right_ips[0] / sample_rate
            
            # 上升时间/速率
            rise_threshold = 1 + 0.1 * peak_amplitude
            rise_start_idx = np.where(f_f0_smooth[:peak_idx] <= rise_threshold)[0][-1] if len(np.where(f_f0_smooth[:peak_idx] <= rise_threshold)[0]) > 0 else 0
            rise_time = (peak_idx - rise_start_idx) / sample_rate
            rise_rate = peak_amplitude / rise_time if rise_time > 0 else np.nan
            
            # 下降时间/速率
            fall_threshold = 1 + 0.1 * peak_amplitude
            fall_end_idx = np.where(f_f0_smooth[peak_idx:] <= fall_threshold)[0][0] + peak_idx if len(np.where(f_f0_smooth[peak_idx:] <= fall_threshold)[0]) > 0 else len(f_f0_smooth)-1
            fall_time = (fall_end_idx - peak_idx) / sample_rate
            fall_rate = peak_amplitude / fall_time if fall_time > 0 else np.nan
            
            # 保存单峰特征
            peak_features.append({
                '峰编号': i+1,
                '峰值时间(秒)': round(peak_time, 3),
                '峰值幅度(ΔF/F0)': round(peak_amplitude, 6),
                '半高宽(秒)': round(fwhm, 3),
                '半高宽左边界(秒)': round(fwhm_left, 3),
                '半高宽右边界(秒)': round(fwhm_right, 3),
                '10%-90%上升时间(秒)': round(rise_time, 3),
                '平均上升速率(ΔF/F0/秒)': round(rise_rate, 6),
                '90%-10%下降时间(秒)': round(fall_time, 3),
                '平均下降速率(ΔF/F0/秒)': round(fall_rate, 6)
            })
        df_peak_features = pd.DataFrame(peak_features)

        # 整体统计特征提取
        overall_stats = {
            '文件名': file_name,
            '总时长(秒)': round(time[-1] - time[0], 3),
            '采样频率(Hz)': round(sample_rate, 2),
            '总数据点数': len(time),
            'F/F0基线值': round(f0_baseline, 6),
            'F/F0全时段均值': round(np.mean(f_f0_smooth), 6),
            'F/F0全时段标准差': round(np.std(f_f0_smooth), 6),
            'F/F0最大值': round(np.max(f_f0_smooth), 6),
            'F/F0最小值': round(np.min(f_f0_smooth), 6),
            'dF/F0全时段均值': round(np.mean(df_f0_smooth), 6),
            'dF/F0全时段标准差': round(np.std(df_f0_smooth), 6),
            '检测到的钙峰总数': peak_count,
            '钙峰平均幅度(ΔF/F0)': round(df_peak_features['峰值幅度(ΔF/F0)'].mean(), 6) if peak_count > 0 else np.nan,
            '钙峰平均半高宽(秒)': round(df_peak_features['半高宽(秒)'].mean(), 3) if peak_count > 0 else np.nan,
            '钙峰平均上升时间(秒)': round(df_peak_features['10%-90%上升时间(秒)'].mean(), 3) if peak_count > 0 else np.nan,
            '钙峰平均下降时间(秒)': round(df_peak_features['90%-10%下降时间(秒)'].mean(), 3) if peak_count > 0 else np.nan,
            '最大钙峰幅度(ΔF/F0)': round(df_peak_features['峰值幅度(ΔF/F0)'].max(), 6) if peak_count > 0 else np.nan,
            '最大钙峰出现时间(秒)': df_peak_features.loc[df_peak_features['峰值幅度(ΔF/F0)'].idxmax(), '峰值时间(秒)'] if peak_count > 0 else np.nan
        }

        # 4.4 生成单文件Excel结果
        single_file_output_path = os.path.join(single_file_output_folder, f"处理结果_{file_name}")
        with pd.ExcelWriter(single_file_output_path, engine='openpyxl') as writer:
            df_raw.to_excel(writer, sheet_name='原始数据', index=False)
            df_preprocess.to_excel(writer, sheet_name='预处理后数据', index=False)
            df_peak_features.to_excel(writer, sheet_name='峰值特征', index=False)
            pd.DataFrame(list(overall_stats.items()), columns=['指标名称', '指标值']).to_excel(writer, sheet_name='整体统计特征', index=False)

        # 美化Excel格式
        wb = load_workbook(single_file_output_path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # 表头样式
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
            # 数据行样式与斑马纹
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = center_alignment
                    cell.border = border
                    if row % 2 == 0:
                        cell.fill = zebra_fill
            # 自动调整列宽
            for col in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col)
                ws.column_dimensions[col_letter].auto_size = True
            # 冻结表头
            ws.freeze_panes = 'A2'
        wb.save(single_file_output_path)

        # 4.5 生成波形图（可选）
        if generate_plot:
            fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(16, 10), sharex=True)
            # F/F0波形
            ax1.plot(time, f_f0_raw, color='gray', alpha=0.5, label='原始F/F0')
            ax1.plot(time, f_f0_smooth, color='#1f77b4', linewidth=1.5, label='预处理后F/F0')
            if peak_count > 0:
                ax1.plot(time[peaks], f_f0_smooth[peaks], 'r*', markersize=10, label='检测到的钙峰')
            ax1.set_ylabel('F/F0', fontsize=12)
            ax1.set_title(f'{file_name} 钙离子成像F/F0波形预处理结果', fontsize=14, fontweight='bold')
            ax1.legend(fontsize=10)
            ax1.grid(alpha=0.3)
            # dF/F0波形
            ax2.plot(time, df_f0_raw, color='gray', alpha=0.5, label='原始dF/F0')
            ax2.plot(time, df_f0_smooth, color='#ff7f0e', linewidth=1.5, label='预处理后dF/F0')
            ax2.set_xlabel('时间(秒)', fontsize=12)
            ax2.set_ylabel('dF/F0', fontsize=12)
            ax2.set_title(f'{file_name} 钙离子成像dF/F0波形预处理结果', fontsize=14, fontweight='bold')
            ax2.legend(fontsize=10)
            ax2.grid(alpha=0.3)
            plt.tight_layout()
            # 保存图片
            plot_file_name = f"{os.path.splitext(file_name)[0]}_波形图.png"
            plt.savefig(os.path.join(plot_output_folder, plot_file_name), dpi=150, bbox_inches='tight')
            plt.close()

        # 返回整体统计特征，用于汇总
        return overall_stats

    except Exception as e:
        print(f"❌  文件 {file_name} 处理失败，错误信息：{str(e)}")
        return None

# ===================== 5. 批量遍历所有文件，执行处理 =====================
# 获取所有符合条件的xlsx文件
file_list = []
for file_name in os.listdir(input_folder):
    # 只处理xlsx文件，且文件名包含指定关键词
    if file_name.endswith('.xlsx') and file_keyword in file_name:
        file_list.append((os.path.join(input_folder, file_name), file_name))

print(f"📁 共找到 {len(file_list)} 个符合条件的xlsx文件，开始批量处理...")

# 遍历所有文件，带进度条
for file_path, file_name in tqdm(file_list, desc="处理进度", unit="文件"):
    file_stats = process_single_file(file_path, file_name)
    if file_stats is not None:
        summary_result_list.append(file_stats)

# ===================== 6. 生成批量处理汇总总表 =====================
if len(summary_result_list) > 0:
    df_summary = pd.DataFrame(summary_result_list)
    summary_output_path = os.path.join(output_folder, "批量处理汇总总表.xlsx")
    
    # 写入汇总表并美化
    with pd.ExcelWriter(summary_output_path, engine='openpyxl') as writer:
        df_summary.to_excel(writer, sheet_name='所有文件汇总结果', index=False)
    
    # 美化汇总表格式
    wb = load_workbook(summary_output_path)
    ws = wb['所有文件汇总结果']
    # 表头样式
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    # 数据行样式与斑马纹
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = center_alignment
            cell.border = border
            if row % 2 == 0:
                cell.fill = zebra_fill
    # 自动调整列宽
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].auto_size = True
    # 冻结表头
    ws.freeze_panes = 'A2'
    wb.save(summary_output_path)

    print(f"✅ 批量处理完成！共成功处理 {len(summary_result_list)} 个文件")
    print(f"📊 单文件处理结果已保存至：{single_file_output_folder}")
    if generate_plot:
        print(f"🖼️  单文件波形图已保存至：{plot_output_folder}")
    print(f"📋  批量汇总总表已保存至：{summary_output_path}")
else:
    print("⚠️  没有成功处理任何文件，请检查文件路径、列名是否匹配")